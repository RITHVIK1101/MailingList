import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from jinja2 import Template


def read_excel_data(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        to, cc, grade, date, name, author, attachments = row
        data.append({'to': to, 'cc': cc, 'grade': grade, 'name': name, 'author': author, 'date': date, 'attachments': attachments})

    return data



def substitute_data(template_path, data):
    with open(template_path, 'r') as file:
        template_content = file.read()

    template = Template(template_content)
    return template.render(data)


def send_email(smtp_server, smtp_port, sender_email, sender_password, to, subject, body, attachments=None):
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = to
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    if attachments:
        for attachment in attachments:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(open(attachment, 'rb').read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment', filename=attachment)
            msg.attach(part)

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)


def main():
    smtp_server = 'smtp.gmail.com'  
    smtp_port = 587 
    sender_email = 'ehscscouncil@gmail.com'  
    sender_password = '0nc3up0n4t1m3' 

    excel_file_path = '/Users/rithviksaba/Documents/test.xlsx' 
    data = read_excel_data(excel_file_path)

    template_path = '/Users/rithviksaba/Documents/test.txt'  
    for item in data:
        substituted_script = substitute_data(template_path, item)
        to = item['to'] 

        subject = 'Your Subject' 
        attachments = [] 
        send_email(smtp_server, smtp_port, sender_email, sender_password, to, subject, substituted_script, attachments)
        print(f'Email sent to {to}')


if __name__ == '__main__':
    main()
